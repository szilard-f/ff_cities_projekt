import os
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
import sqlite3
import re
from datetime import datetime
from cryptography.fernet import Fernet
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter

# Hálózati adatbázis könyvtár
DB_DIR = r"W:\depo10\ffcities"
DB_FILE = os.path.join(DB_DIR, "data.db")
DB_ENC_FILE = os.path.join(DB_DIR, "data.db.enc")

# Várt oszlopnevek 
COLUMNS = ["depószám", "depónév", "irányítószám", "helység", "nap", "nap +", "megjegyzés", "orig_order"]

# Oszlopszélességek 
COL_WIDTH_SORSZAM = 40
COL_WIDTH_DEPOSZAM = 50
COL_WIDTH_DEPONEV = 140
COL_WIDTH_IRANYITOSZAM = 60
COL_WIDTH_HELYSEG = 200
COL_WIDTH_NAP = 150
COL_WIDTH_NAP2 = 100
COL_WIDTH_MEGJEGYZES = 150

# fix titkosítási kulcs 
KEY = b'KXCEpLXa4cL5_gPFvG4R1Bfyl-cRtLQmyJ8iOD1UrLA='
fernet = Fernet(KEY)

class PasswordDialog(tk.Toplevel):
    def __init__(self, master, correct_password="4422"):
        super().__init__(master)
        self.title("Jelszó szükséges")
        self.geometry("300x150")
        self.resizable(False, False)
        self.correct_password = correct_password
        self.password = None
        self.center_window(300, 150)
        
        tk.Label(self, text="Add meg a jelszót:", font=("Arial", 12)).pack(pady=10)
        self.entry = tk.Entry(self, show="*", font=("Arial", 12))
        self.entry.pack(pady=5)
        self.entry.focus()
        tk.Button(self, text="OK", command=self.check_password, font=("Arial", 10)).pack(pady=10)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.wait_window(self)

    def center_window(self, width, height):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")

    def check_password(self):
        if self.entry.get() == self.correct_password:
            self.password = True
            self.destroy()
        else:
            messagebox.showerror("Hiba", "Helytelen jelszó!")
            self.lift()
            self.focus_force()
            self.entry.focus_set()
            self.entry.delete(0, tk.END)

    def on_close(self):
        self.password = False
        self.destroy()

def check_password(parent=None):
    if parent is None:
        parent = tk._default_root  # Ha nincs megadva, akkor az aktuális főablakot használja
    dialog = PasswordDialog(parent)
    return dialog.password

class DataApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Family Frost - település lista")
        self.geometry("900x600")
        self.center_window()
        self.create_widgets()
        self.load_from_database()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def center_window(self):
        win_width = 1200
        win_height = 600
        screen_width = self.winfo_screenwidth()
        x = (screen_width - win_width) // 2
        y = 40
        self.geometry(f"{win_width}x{win_height}+{x}+{y}")

    def create_widgets(self):
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="Import", command=self.import_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Összes export", command=self.export_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Depó törlése", command=self.delete_depot).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Teljes törlés", command=self.delete_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Kilépés", command=self.on_closing).pack(side=tk.LEFT, padx=5)

    def import_file(self):
        if not check_password(self):
            return
        messagebox.showinfo("Import", "Importálás sikeres!")

    def export_all(self):
        if not check_password():
            return
        messagebox.showinfo("Export", "Összes adat exportálása kész!")

    def delete_depot(self):
        if not check_password():
            return
        messagebox.showinfo("Törlés", "Depó törölve!")

    def delete_all(self):
        if not check_password():
            return
        messagebox.showinfo("Törlés", "Minden adat törölve!")

    def load_from_database(self):
        pass  # Itt kerül be a valódi adatbázis betöltés logika

    def on_closing(self):
        if messagebox.askyesno("Kilépés", "Biztosan be akarod zárni az alkalmazást?"):
            self.quit()
            self.destroy()









def clean_text(text):
    """Szöveg tisztít"""
    if isinstance(text, str):
        return re.sub(r'\s+', ' ', text).strip()
    return text

def day_order(day):
    """Egyedi rendezési sorrend a napokhoz."""
    mapping = {"hétfő": 1, "kedd": 2, "szerda": 3, "csütörtök": 4, "péntek": 5, "szombat": 6, "vasárnap": 7}
    if isinstance(day, str):
        return mapping.get(day.lower().strip(), 999)
    return 999



def process_irányítószám(val):
    
    if pd.isnull(val):
        return ""
    try:
        if isinstance(val, (int, float)):
            if float(val).is_integer():
                return str(int(val))
            else:
                return str(val).strip()
        val_str = clean_text(val)
        try:
            num = float(val_str)
            if num.is_integer():
                return str(int(num))
            else:
                return val_str
        except:
            return val_str
    except Exception:
        return str(val).strip()

def convert_to_number_if_possible(val):
    
    try:
        if isinstance(val, str) and ',' in val:
            return val
        num = float(val)
        if num.is_integer():
            return int(num)
        else:
            return num
    except:
        return val

def encrypt_file(input_filename, output_filename):
    """Fájl titkosítása."""
    with open(input_filename, 'rb') as f:
        data = f.read()
    encrypted = fernet.encrypt(data)
    with open(output_filename, 'wb') as f:
        f.write(encrypted)

def decrypt_file(input_filename, output_filename):
    """Fájl visszafejtése."""
    with open(input_filename, 'rb') as f:
        data = f.read()
    decrypted = fernet.decrypt(data)
    with open(output_filename, 'wb') as f:
        f.write(decrypted)


def protect_headers(filename):
    wb = load_workbook(filename)
    ws = wb.active

    # Először minden cellát zároljuk alapból (Excel alapértelmezés)
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)  # Minden cella zárolva van

    # Csak az első sort hagyjuk védetten (fejléc)
    for cell in ws[1]:
        cell.protection = Protection(locked=True)  # Fejléc védett marad

    # Az A2:G10000 tartomány feloldása a védelem alól
    for row in ws.iter_rows(min_row=2, max_row=4000, min_col=1, max_col=7):
        for cell in row:
            cell.protection = Protection(locked=False)  # Szerkeszthetővé tesszük

    # Munkalapvédelem bekapcsolása
    ws.protection.sheet = True
    ws.protection.enable()

    # Oszlopszélesség automatikus beállítása
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(filename)


def get_unique_key(row):
    
    return (
        str(row['depószám']).strip(),
        str(row['depónév']).strip().lower(),
        str(row['irányítószám']).strip(),
        str(row['helység']).strip().lower(),
        #str(row['nap']).strip().lower(),
        #str(row['nap +']).strip().lower()
    )

class DataApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Family Frost - település lista")
        self.geometry("900x600")
        self.center_top()
        self.data = pd.DataFrame(columns=COLUMNS)
        self.sort_order = {}
        self.last_clicked_column = None
        self.last_click_time = 0
        self.create_widgets()
        self.load_from_database()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def center_top(self):
        win_width = 1200
        win_height = 600
        screen_width = self.winfo_screenwidth()
        x = (screen_width - win_width) // 2
        y = 40
        self.geometry(f"{win_width}x{win_height}+{x}+{y}")

    def create_widgets(self):
        # Keresőmezők
        search_frame = ttk.Frame(self)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(search_frame, text="Depószám:").pack(side=tk.LEFT, padx=5)
        self.depo_entry = ttk.Entry(search_frame, width=10)
        self.depo_entry.pack(side=tk.LEFT, padx=5)
        self.depo_entry.bind("<KeyRelease>", self.search_data)
        ttk.Label(search_frame, text="Irányítószám:").pack(side=tk.LEFT, padx=5)
        self.postal_entry = ttk.Entry(search_frame, width=15)
        self.postal_entry.pack(side=tk.LEFT, padx=5)
        self.postal_entry.bind("<KeyRelease>", self.search_data)
        ttk.Label(search_frame, text="Helység:").pack(side=tk.LEFT, padx=5)
        self.helység_entry = ttk.Entry(search_frame, width=20)
        self.helység_entry.pack(side=tk.LEFT, padx=5)
        self.helység_entry.bind("<KeyRelease>", self.search_data)
        clear_filter_btn = ttk.Button(search_frame, text="Szűrés törlése", command=self.clear_filter)
        clear_filter_btn.pack(side=tk.LEFT, padx=5)

        # Treeview – oszlopok: sorszám, depószám, depónév, irányítószám, helység, nap, nap +, megjegyzés
        self.tree_frame = ttk.Frame(self)
        self.tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.tree_left_frame = ttk.Frame(self.tree_frame)
        self.tree_left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        columns = ("sorszám", "depószám", "depónév", "irányítószám", "helység", "nap", "nap +", "megjegyzés")
        self.tree = ttk.Treeview(self.tree_left_frame, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
            if col == "sorszám":
                self.tree.column(col, width=COL_WIDTH_SORSZAM, anchor="center")
            elif col == "depószám":
                self.tree.column(col, width=COL_WIDTH_DEPOSZAM, anchor="center")
            elif col == "depónév":
                self.tree.column(col, width=COL_WIDTH_DEPONEV, anchor="center")
            elif col == "irányítószám":
                self.tree.column(col, width=COL_WIDTH_IRANYITOSZAM, anchor="center")
            elif col == "helység":
                self.tree.column(col, width=COL_WIDTH_HELYSEG, anchor="w")
            elif col == "nap":
                self.tree.column(col, width=COL_WIDTH_NAP, anchor="center")
            elif col == "nap +":
                self.tree.column(col, width=COL_WIDTH_NAP2, anchor="center")
            elif col == "megjegyzés":
                self.tree.column(col, width=COL_WIDTH_MEGJEGYZES, anchor="center")
            else:
                self.tree.column(col, width=100, anchor="center")
        self.tree.pack(fill=tk.BOTH, expand=True)
        #self.tree.bind("<Double-1>", lambda e: self.edit_record())

        # Jobb oldali rész: scrollbar, jump gombok
        self.right_frame = ttk.Frame(self.tree_frame)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.v_scrollbar = ttk.Scrollbar(self.right_frame, orient="vertical", command=self.tree.yview)
        self.v_scrollbar.pack(side=tk.TOP, fill=tk.Y, expand=True)
        self.tree.configure(yscrollcommand=self.v_scrollbar.set)
        self.jump_frame = ttk.Frame(self.right_frame)
        self.jump_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        #top_btn = ttk.Button(self.jump_frame, text="Ugrás tetejére", command=self.scroll_to_top)
        #top_btn.pack(fill=tk.X, pady=2)
        #bottom_btn = ttk.Button(self.jump_frame, text="Ugrás aljára", command=self.scroll_to_bottom)
        #bottom_btn.pack(fill=tk.X, pady=2)

        # Alsó gombpanel: Import, Szerkesztés, Export depónként, Összes export, Depó törlése, Teljes törlés, Kilépés
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=10)
        import_button = ttk.Button(button_frame, text="Import", command=self.import_file)
        import_button.pack(side=tk.LEFT, padx=5)
        edit_button = ttk.Button(button_frame, text="Szerkesztés", command=self.edit_record)
        edit_button.pack(side=tk.LEFT, padx=5)
        export_depot_button = ttk.Button(button_frame, text="Export depónként", command=self.export_by_depot)
        export_depot_button.pack(side=tk.LEFT, padx=5)
        export_all_button = ttk.Button(button_frame, text="Összes export", command=self.export_all)
        export_all_button.pack(side=tk.LEFT, padx=5)
        delete_depot_button = ttk.Button(button_frame, text="Depó törlése", command=self.delete_depot)
        delete_depot_button.pack(side=tk.LEFT, padx=5)
        delete_all_button = ttk.Button(button_frame, text="Teljes törlés", command=self.delete_all)
        delete_all_button.pack(side=tk.LEFT, padx=5)
        exit_button = ttk.Button(button_frame, text="Kilépés", command=self.on_closing)
        exit_button.pack(side=tk.LEFT, padx=5)

    def export_all(self):
        """Összes adatot egyetlen Excel fájlba"""
        if not check_password():
            return
        if self.data.empty:
            messagebox.showinfo("Info", "Nincs exportálható adat.")
            return
        try:
            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"FF_településlista_{date_str}.xlsx"
            export_df = self.data.copy()
            export_df = export_df[["depószám", "depónév", "irányítószám", "helység", "nap", "nap +", "megjegyzés"]]
            export_df["irányítószám"] = export_df["irányítószám"].apply(lambda x: process_irányítószám(x))
            export_df.to_excel(filename, index=False, engine='openpyxl')
            protect_headers(filename)
            messagebox.showinfo("Export", f"Összes exportálás kész:\n{filename}")
        except Exception as e:
            messagebox.showerror("Hiba", f"Összes exportálás hiba: {e}")

    def clear_filter(self):
        self.depo_entry.delete(0, tk.END)
        self.postal_entry.delete(0, tk.END)
        self.helység_entry.delete(0, tk.END)
        self.refresh_treeview(self.data)

    def clear_filter_fields(self):
        self.depo_entry.delete(0, tk.END)
        self.postal_entry.delete(0, tk.END)
        self.helység_entry.delete(0, tk.END)

    def load_from_database(self):
        if os.path.exists(DB_ENC_FILE):
            try:
                decrypt_file(DB_ENC_FILE, DB_FILE)
            except Exception as e:
                messagebox.showerror("Hiba", f"Adatbázis visszafejtési hiba: {e}")
        if os.path.exists(DB_FILE):
            try:
                conn = sqlite3.connect(DB_FILE)
                try:
                    df = pd.read_sql_query("SELECT * FROM munkafüzet", conn, index_col=None)
                except Exception as e:
                    if "no such table" in str(e).lower():
                        df = pd.DataFrame(columns=COLUMNS)
                    else:
                        raise e
                conn.close()
                self.data = df
                self.refresh_treeview(self.data)
            except Exception as e:
                messagebox.showerror("Hiba", f"Hiba az adatbázis betöltésekor: {e}")

    def scroll_to_top(self):
        self.tree.yview_moveto(0)

    def scroll_to_bottom(self):
        self.tree.yview_moveto(1)

    def import_file(self):
        if not check_password():
            return
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])
        if file_path:
            try:
                new_df = pd.read_excel(file_path)
                new_df.columns = [col.lower() for col in new_df.columns]
                required_cols = ['depószám', 'depónév', 'irányítószám', 'helység', 'nap', 'nap +']
                for req in required_cols:
                    if req not in new_df.columns:
                        messagebox.showerror("Hiba", f"Hiányzó oszlop: {req}")
                        return
                if 'megjegyzés' not in new_df.columns:
                    new_df['megjegyzés'] = ""
                for col in new_df.columns:
                    if new_df[col].dtype == 'object':
                        new_df[col] = new_df[col].apply(clean_text)
                new_df['depószám'] = new_df['depószám'].apply(lambda x: int(x) if pd.notnull(x) else x)
                new_df['irányítószám'] = new_df['irányítószám'].apply(lambda x: process_irányítószám(x))
                new_df['nap'] = new_df['nap'].apply(clean_text)
                new_df['nap +'] = new_df['nap +'].fillna("")
                new_df['megjegyzés'] = new_df['megjegyzés'].fillna("")
                new_df['orig_order'] = range(1, len(new_df) + 1)
                new_df.sort_values(by='helység', inplace=True)
                
                updated_count = 0
                new_count = 0
                if self.data.empty:
                    self.data = new_df.copy()
                    new_count = len(new_df)
                else:
                    for idx, new_row in new_df.iterrows():
                        key = get_unique_key(new_row)
                        mask = self.data.apply(lambda row: get_unique_key(row) == key, axis=1)
                        if mask.any():
                            match_index = self.data[mask].index[0]
                            self.data.loc[match_index] = new_row
                            updated_count += 1
                        else:
                            self.data = pd.concat([self.data, new_row.to_frame().T], ignore_index=True)
                            new_count += 1
                conn = sqlite3.connect(DB_FILE)
                self.data.to_sql("munkafüzet", conn, if_exists="replace", index=False)
                conn.close()
                self.refresh_treeview(self.data)
                messagebox.showinfo("Import", f"Importálás kész.\nFrissített rekordok: {updated_count}\nÚj rekordok: {new_count}")
            except Exception as e:
                messagebox.showerror("Hiba", f"Hiba a fájl importálása során: {e}")

    def refresh_treeview(self, df):
        self.tree.delete(*self.tree.get_children())
        for sorszam, (orig_index, row) in enumerate(df.iterrows(), start=1):
            dep_szam = int(row['depószám']) if pd.notnull(row['depószám']) else ""
            nap2 = row['nap +'] if pd.notna(row['nap +']) else ""
            values = (sorszam,
                      dep_szam,
                      row['depónév'],
                      row['irányítószám'],
                      row['helység'],
                      row['nap'],
                      nap2,
                      row.get("megjegyzés", ""))
            self.tree.insert("", "end", iid=str(orig_index), values=values)

    def sort_by_column(self, col):
        current_time = time.time()
        double_click = False
        if hasattr(self, "last_clicked_column") and self.last_clicked_column == col:
            if current_time - self.last_click_time < 0.5:
                double_click = True
        self.last_clicked_column = col
        self.last_click_time = current_time
        ascending = self.sort_order.get(col, True)
        if double_click or col == "depószám":
            # Double click vagy a primary oszlop: kizárólag az adott oszlop szerint rendezünk.
            if col in ("nap", "nap +"):
                self.data.sort_values(by=col, key=lambda s: s.apply(day_order), ascending=ascending, inplace=True, kind='mergesort')
            elif col == "irányítószám":
                self.data.sort_values(by=col, key=lambda s: s.astype(str), ascending=ascending, inplace=True, kind='mergesort')
            else:
                self.data.sort_values(by=col, ascending=ascending, inplace=True, kind='mergesort')
        else:
            
            if col in ("nap", "nap +"):
                self.data.sort_values(by=col, key=lambda s: s.apply(day_order), ascending=ascending, inplace=True, kind='mergesort')
            elif col == "irányítószám":
                self.data.sort_values(by=col, key=lambda s: s.astype(str), ascending=ascending, inplace=True, kind='mergesort')
            else:
                self.data.sort_values(by=col, ascending=ascending, inplace=True, kind='mergesort')
            self.data.sort_values(by="depószám", ascending=True, inplace=True, kind='mergesort')
        self.sort_order[col] = not ascending
        self.refresh_treeview(self.data)

    def search_data(self, event=None):
        depo_filter = self.depo_entry.get().strip()
        postal = self.postal_entry.get().strip()
        helyseg = self.helység_entry.get().strip().lower()
        df_filtered = self.data.copy()
        if depo_filter:
            df_filtered = df_filtered[df_filtered['depószám'].astype(str).str.contains(depo_filter)]
        if postal:
            df_filtered = df_filtered[df_filtered['irányítószám'].astype(str).str.contains(postal)]
        if helyseg:
            df_filtered = df_filtered[df_filtered['helység'].astype(str).str.lower().str.contains(helyseg)]
        print("Találatok:")
        print(df_filtered)
        self.refresh_treeview(df_filtered)

    def edit_record(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Info", "Kérem válasszon ki egy rekordot a szerkesztéshez.")
            return
        item_id = selected[0]
        try:
            record = self.data.loc[int(item_id)]
        except Exception:
            record = self.data.loc[item_id]
        EditWindow(self, record, item_id)

    def update_record(self, index, updated_record):
        for key, val in updated_record.items():
            new_val = val
            if isinstance(new_val, str):
                new_val = clean_text(new_val)
            if key == 'depószám':
                try:
                    new_val = int(float(new_val))
                except ValueError:
                    new_val = None
            elif key == 'irányítószám':
                new_val = process_irányítószám(new_val)
            elif key in ("nap", "nap +"):
                new_val = clean_text(new_val)
            self.data.at[int(index), key] = new_val
        self.refresh_treeview(self.data)
        self.clear_filter_fields()

    def delete_depot(self):
        if not check_password():
            return
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Info", "Kérem válasszon ki egy rekordot (depószám alapján) a törléshez.")
            return
        item_id = selected[0]
        try:
            row = self.data.loc[int(item_id)]
        except Exception:
            row = self.data.loc[item_id]
        depo_to_delete = row['depószám']
        confirm = messagebox.askyesno("Megerősítés", f"Biztosan törlöd a depó {depo_to_delete} összes rekordját?")
        if confirm:
            self.data = self.data[self.data['depószám'] != depo_to_delete]
            if self.data.empty:
                if os.path.exists(DB_FILE):
                    os.remove(DB_FILE)
                if os.path.exists(DB_ENC_FILE):
                    os.remove(DB_ENC_FILE)
            else:
                conn = sqlite3.connect(DB_FILE)
                self.data.to_sql("munkafüzet", conn, if_exists="replace", index=False)
                conn.close()
            self.refresh_treeview(self.data)

    def delete_all(self):
        if not check_password():
            return
        confirm = messagebox.askyesno("Megerősítés", "Figyelem: ez minden adatot töröl! Biztosan törlöd az egész adatbázist?")
        if confirm:
            self.data = pd.DataFrame(columns=COLUMNS)
            if os.path.exists(DB_FILE):
                os.remove(DB_FILE)
            if os.path.exists(DB_ENC_FILE):
                os.remove(DB_ENC_FILE)
            self.refresh_treeview(self.data)







    def export_by_depot(self):
        if not check_password():
            return
        if self.data.empty:
            messagebox.showinfo("Info", "Nincs exportálható adat.")
            return
        grouped = self.data.groupby("depószám")
        export_count = 0
        for depo, group in grouped:
            if pd.isnull(depo):
                continue
            try:
                depo_int = int(float(depo))
            except Exception:
                depo_int = depo
            filename_depo = f"{depo_int:02d}"
            deponev = group.iloc[0]['depónév']
            date_str = datetime.now().strftime("%Y%m%d")
            safe_depot_name = re.sub(r'[^a-zA-Z0-9áéíóöőúüűÁÉÍÓÖŐÚÜŰ]', '_', deponev)
            safe_depot_name = re.sub(r'_+', '_', safe_depot_name).strip('_')
            filename = f"{filename_depo}_{safe_depot_name}_településlista_{date_str}.xlsx"
            try:
                export_df = group.copy()
                export_df = export_df[["depószám", "depónév", "irányítószám", "helység", "nap", "nap +", "megjegyzés"]]
                export_df["irányítószám"] = export_df["irányítószám"].apply(lambda x: process_irányítószám(x))
                export_df["irányítószám"] = pd.to_numeric(export_df["irányítószám"], errors='coerce')
                export_df.to_excel(filename, index=False, engine='openpyxl')
                protect_headers(filename)
                export_count += 1
            except Exception as e:
                messagebox.showerror("Hiba", f"Export hiba ({filename}): {e}")
                return
        messagebox.showinfo("Export", f"Exportálás kész. {export_count} fájl lett mentve.")





    def on_closing(self):
        if self.data.empty:
            self.destroy()
            return
        try:
            conn = sqlite3.connect(DB_FILE)
            self.data.to_sql("munkafüzet", conn, if_exists="replace", index=False)
            conn.close()
            encrypt_file(DB_FILE, DB_ENC_FILE)
            os.remove(DB_FILE)
        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba a mentés és titkosítás során: {e}")
        self.destroy()

import tkinter as tk

class EditWindow(tk.Toplevel):
    def __init__(self, master, record, record_index):
        super().__init__(master)
        self.title("Rekord szerkesztése")
        self.master = master
        self.record_index = record_index
        self.record = record

        self.create_widgets()
        
        # Ablak középre helyezése
        self.center_window(260, 280)

    def center_window(self, width, height):
        """Ablakot a képernyő közepére igazítja."""
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        self.resizable(False, False)
        
        

    def create_widgets(self):
        fields = ['depószám', 'depónév', 'irányítószám', 'helység', 'nap', 'nap +', 'megjegyzés']
        self.entries = {}
        for i, field in enumerate(fields):
            ttk.Label(self, text=f"{field}:").grid(row=i, column=0, padx=5, pady=5, sticky=tk.W)
            entry = ttk.Entry(self, width=25)
            entry.grid(row=i, column=1, padx=5, pady=5)
            value = self.record[field] if field in self.record else ""
            if field == 'depószám':
                try:
                    f_val = float(value)
                    if f_val.is_integer():
                        value = int(f_val)
                except Exception:
                    pass
            entry.insert(0, str(value))
            self.entries[field] = entry
        button_frame = ttk.Frame(self)
        button_frame.grid(row=len(fields), column=0, columnspan=2, pady=10)
        save_button = ttk.Button(button_frame, text="Mentés", command=self.save)
        save_button.pack(side=tk.LEFT, padx=5)
        exit_button = ttk.Button(button_frame, text="Kilépés", command=self.destroy)
        exit_button.pack(side=tk.LEFT, padx=5)

    def save(self):
        updated = {}
        for field, entry in self.entries.items():
            updated[field] = entry.get()
        self.master.update_record(self.record_index, updated)
        self.destroy()

if __name__ == "__main__":
    app = DataApp()
    app.mainloop()

